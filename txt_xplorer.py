'''Code is written by: Sumit Arjun.
This code is offline working one. 
The code is sample code for data extraction of institution data from pre-collected data.Just kinda sorter. 
Do make changes according to the information of data that is asked!!
I am not responsible if your boss kicks you out, if inappropriate data is served to him..☠☠☠
Edit/Debug count: 21'''

import re
import openpyxl

# Function to extract details from a single entry
def extract_details(entry):
    details = {}
    
    # Regular expressions for each detail. Regular expression depends on the data you are searching... 
    school_name_re = re.compile(r'Name:\s*(.+?)(?=\n|$)')
    head_name_re = re.compile(r'Head/Principal Name:\s*(.+?)(?=\n|$)')
    status_re = re.compile(r'Status of the School:\s*(.+?)(?=\n|$)')
    phone_re = re.compile(r'Phone No:\s*(.+?)(?=\n|$)')
    email_re = re.compile(r'Email:\s*(.+?)(?=\n|$)')
    
    details['School Name'] = school_name_re.search(entry).group(1) if school_name_re.search(entry) else ''
    details['Head/Principal Name'] = head_name_re.search(entry).group(1) if head_name_re.search(entry) else ''
    details['Status of School'] = status_re.search(entry).group(1) if status_re.search(entry) else ''
    details['Phone No'] = phone_re.search(entry).group(1) if phone_re.search(entry) else ''
    details['Email'] = email_re.search(entry).group(1) if email_re.search(entry) else ''
    
    return details

# Prompt the user for the file path of the text file. pls provide better format of data. 
file_path = input("Please enter the path of the text file: ")

# Read the data from the text file
with open(file_path, 'r') as file:
    data = file.read()

# Assuming each school's data is separated by a line starting with a number
entries = re.split(r'\n(?=\d)', data.strip())

# Extract details for each entry
school_details = [extract_details(entry) for entry in entries if entry.strip()]

# Create a new Excel workbook and select the active worksheet
workbook = openpyxl.Workbook()
sheet = workbook.active
sheet.title = 'Details'

# Define the headers
headers = ['School Name', 'Head/Principal Name', 'Status of School', 'Phone No', 'Email']

# Write the headers to the worksheet
for col_num, header in enumerate(headers, 1):
    sheet.cell(row=1, column=col_num, value=header)

# Write the extracted details to the worksheet
for row_num, details in enumerate(school_details, 2):
    for col_num, header in enumerate(headers, 1):
        sheet.cell(row=row_num, column=col_num, value=details.get(header, ''))

# Save the workbook
workbook.save('school_details.xlsx')

print("Data has been successfully extracted and saved to 'school_details.xlsx'.")
